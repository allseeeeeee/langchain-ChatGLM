import json

if __name__ == '__main__':
    filename = r"/home/dev/team/langchain-ChatChat/knowledge_base/幼儿信息/content/1711012368266.xlsx"
    from unstructured.partition.xlsx import partition_xlsx
    result = partition_xlsx(filename=filename)
    print(result)
    import openpyxl, requests, os
    wb = openpyxl.open(filename=filename, read_only=True)
    ws = wb[wb.sheetnames[0]]
    titles = []
    docs = []
    for i in range(ws.max_row):
        row = {}
        for j in range(ws.max_column):
            value = ws.cell(i + 1, j + 1).value
            if i == 0:
                titles.append(value)
            else:
                row[titles[j]] = value
                print(f"{i}行{j}列{titles[j]}值为{value}")
        if i != 0 and row:

            docs.append({"page_content": f"{row[titles[0]]}#{row[titles[1]]}#{row[titles[2]]}#{row[titles[3]]}", "metadata": {}, "type": "Document"})
    print(json.dumps({os.path.basename(filename): docs}, ensure_ascii=False))

    host = "http://192.168.0.182:17861"

    # req = {
    #     "knowledge_base_name": "班级信息",
    #     "vector_store_type": "faiss",
    #     "embed_model": "bge-large-zh-v1.5",
    # }
    # resp = requests.post(f"{host}/knowledge_base/create_knowledge_base", json=req, headers={"content-type": "application/json"})
    # print("创建知识库", resp.json())
    data = {
        "knowledge_base_name": "班级信息",
        "override": True,
        "to_vector_store": True,
        "chunk_size": 250,
        "chunk_overlap": 50,
        "zh_title_enhance": False,
        "not_refresh_vs_store": False,
        "docs": json.dumps({os.path.basename(filename): docs}, ensure_ascii=False)
    }
    resp = requests.post(f"{host}/knowledge_base/upload_docs", data=data,
                         # headers={"content-type": "multipart/form-data"},
                         files={"files": (os.path.basename(filename), open(filename, 'rb'))})
    print(resp.text)
    pass
